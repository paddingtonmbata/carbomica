"""
Define atomica project based on input data spreadsheet.
"""
import atomica as at
import pandas as pd
from books import generate_books
import streamlit as st
import openpyxl
import numpy as np
from scenarios import coverage_scenario, budget_scenario, optimization
import os

from datetime import datetime

# ## Time frame of simulation
# start_year = 2024 # MODIFY AS NEEDED
# end_year = start_year + 5 # MODIFY AS NEEDED

def delete_books(file_path):
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except OSError as e:
            print(f"Error: {e}")
    else:
        print(f"File '{file_path}' does not exist.")

#need to create our own "input_data_example.xlsx" file based on user input
def create_excel_file(file_path):
    delete_books(file_path)
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    facility_sheet = workbook.create_sheet('facility')
    facility_sheet.append(['Code Name', 'Display Name'])
    interventions = workbook.create_sheet('interventions')
    interventions.append(['Code Name', 'Display Name'])
    emission_sources_sheet = workbook.create_sheet('emission sources')
    emission_sources_sheet.append(['Code Name', 'Display Name'])
    emission_data_sheet = workbook.create_sheet('emission data')
    emission_data_sheet.append(['facilities', 'Grid_Electricity', 'Grid_Gas', 'Bottled_Gas', 'Liquid_Fuel', 'Vehicle_Fuel_Owned', 'Business_Travel', 'Anaesthetic_Gases', 'Refrigeration_Gases', 'Waste_Management', 'Medical_Inhalers'])    
    interventions_sheet = workbook.create_sheet('interventions')
    interventions_sheet.append(['Code Name', 'Display Name'])
    emission_targets_sheet = workbook.create_sheet('emission targets')
    emission_targets_sheet.append(['interventions','Grid_Electricity', 'Grid_Gas', 'Bottled_Gas', 'Liquid_Fuel', 'Vehicle_Fuel_Owned', 'Business_Travel', 'Anaesthetic_Gases', 'Refrigeration_Gases', 'Waste_Management', 'Medical_Inhalers'])
    effect_sizes_sheet = workbook.create_sheet('effect sizes')
    effect_sizes_sheet.append(['facilities', 'Recycling_WasteSegregation_effect', 'SolarSystem_Installation_effect', 'Efficient_Chillers_Upgrade_effect', 'Lighting_Efficiency_effect', 'LowGWP_Refrigerants_effect', 'Hybrid_Car_Use_effect', 'LowGWP_Inhalers_effect', 'LowGWP_AnaestheticGases_effect', 'Staff_Training_Awareness_effect'])
    implementation_costs_sheet = workbook.create_sheet('implementation costs')
    implementation_costs_sheet.append(['facilities', 'Recycling_WasteSegregation_cost', 'SolarSystem_Installation_cost', 'Efficient_Chillers_Upgrade_cost', 'Lighting_Efficiency_cost', 'LowGWP_Refrigerants_cost', 'Hybrid_Car_Use_cost', 'LowGWP_Inhalers_cost', 'LowGWP_AnaestheticGases_cost', 'Staff_Training_Awareness_cost'])
    maintenance_costs_sheet = workbook.create_sheet('maintenance costs')
    maintenance_costs_sheet.append(['facilities', 'Recycling_WasteSegregation_cost', 'SolarSystem_Installation_cost', 'Efficient_Chillers_Upgrade_cost', 'Lighting_Efficiency_cost', 'LowGWP_Refrigerants_cost', 'Hybrid_Car_Use_cost', 'LowGWP_Inhalers_cost', 'LowGWP_AnaestheticGases_cost', 'Staff_Training_Awareness_cost'])
    emission_sources_data = [
        ['Grid_Electricity', 'Grid Electricity'],
        ['Grid_Gas', 'Grid gas'],
        ['Bottled_Gas', 'Bottled gas (LPG)'],
        ['Liquid_Fuel', 'Liquid fuel (Petrol or Diesel)'],
        ['Vehicle_Fuel_Owned', 'Vehicle Fuel (Owned Vehicles)'],
        ['Business_Travel', 'Business travel (Taxi, Car hires, Train, Air travel, Local bus)'],
        ['Anaesthetic_Gases', 'Anaesthetic gases'],
        ['Refrigeration_Gases', 'Refrigerants'],
        ['Waste_Management', 'Waste'],
        ['Medical_Inhalers', 'Inhalers'],
    ]

    # Populate the sheet with constant data
    for row_data in emission_sources_data:
        emission_sources_sheet.append(row_data)    

    emission_targets_data = [
        ['Recycling_WasteSegregation'   ,'','','','','','','','','y',''],
        ['SolarSystem_Installation'     ,'y','','','','','','','','',''],
        ['Efficient_Chillers_Upgrade'   ,'y','','','','','','','y','',''],
        ['Lighting_Efficiency'          ,'y','','','','','','','','',''],
        ['LowGWP_Refrigerants'          ,'','','','','','','','y','',''],
        ['Hybrid_Car_Use'               ,'','','','','y','','','','',''],
        ['LowGWP_Inhalers'              ,'','','','','','','','','','y'],
        ['LowGWP_AnaestheticGases'      ,'','','','','','','y','','',''],
        ['Staff_Training_Awareness'     ,'y','','','','','','','y','y','y'],
    ]

    # Populate the sheet with constant data
    for row_data in emission_targets_data:
        emission_targets_sheet.append(row_data)

    workbook.save(file_path)

def update_sheet(file_path, selected_facility, interventions, emission_data, effect_sizes, implementation_cost, maintenance_cost):
    workbook = openpyxl.load_workbook(file_path)
    facility_sheet = workbook['facility']
    interventions_sheet = workbook['interventions']
    emission_data_sheet = workbook['emission data']
    emission_data_sheet.append(emission_data)
    effect_sizes_sheet = workbook['effect sizes']
    effect_sizes_sheet.append(effect_sizes)
    implementation_costs_sheet = workbook['implementation costs']
    implementation_costs_sheet.append(implementation_cost)
    maintenance_costs_sheet = workbook['maintenance costs']
    maintenance_costs_sheet.append(maintenance_cost)
    formatted_data = []

    for intervention in interventions:
        formatted_name = intervention.replace(' ', '_')
        # Additional formatting for Display Name
        formatted_display_name = intervention.replace('_', ' ')
        formatted_data.append([formatted_name, formatted_display_name])

    # Assuming interventions_sheet is a pandas DataFrame
    for row_data in formatted_data:
        interventions_sheet.append(row_data)
    
    # Clear existing data in the sheet (excluding headers)
    for row in facility_sheet.iter_rows(min_row=2, max_row=facility_sheet.max_row, max_col=facility_sheet.max_column):
        for cell in row:
            cell.value = None    
    # Example data for selected facility (replace with your actual data)
    facility_data = {
        'AKHS_Mombasa': [['AKHS_Mombasa', 'Aga Khan Hospital, Mombasa']],
    }

    # Populate the sheet with data for the selected facility
    for row_data in facility_data.get(selected_facility, []):
        facility_sheet.append(row_data)    

    workbook.save(file_path)

if __name__ == "__main__":
    # Specify the file path where you want to save the Excel file
    file_path = "input_data.xlsx"
    # Create the Excel file initially
    delete_books(file_path)
    create_excel_file(file_path)
    # Streamlit UI
    with st.sidebar:
        (st.title("Facility Data Generator"))

        # Dropdown to select a facility
        selected_facility = st.selectbox("Select Facility", ['AKHS_Mombasa'])
        start_year = st.number_input("Starting year", value=datetime.now().year, max_value=datetime.now().year+10)
        end_year = st.number_input("Ending year", value=datetime.now().year+5, max_value=9999)
        st.header("Intervention")
        interventions = st.multiselect("Select the interventions", [
            'Recycling WasteSegregation', 
            'SolarSystem Installation', 
            'Efficient Chillers Upgrade', 
            'Lighting Efficiency', 
            'LowGWP Refrigerants', 
            'Hybrid Car Use',
            'LowGWP Inhalers',
            'LowGWP AnaestheticGases',
            'Staff Training_Awareness'
        ])
        st.header("Emission data inputs")
        em_data_grid_electricity = st.number_input("Grid Electricity", step=1000, value=157957
)
        em_data_grid_gas = st.number_input("Grid gas", step=1000, value=0)
        em_data_bottled_gas = st.number_input("Bottled gas", step=1000, value=40600)
        em_data_liquid_fuel = st.number_input("Liquid fuel", step=1000, value=7890)
        em_data_vehicle_fuel_owned = st.number_input("Vehicle fuel owned", step=1000, value=8834)
        em_data_business_travel = st.number_input("Business travel", step=1000, value=45472)
        em_data_anaesthetic_gases = st.number_input("Anaesthetic gases", step=1000, value=36604)
        em_data_refrigeration_gases = st.number_input("Refrigeration gases", step=1000, value=69090)
        em_data_waste_management = st.number_input("Waste management", step=1000, value=186383)
        em_data_medical_inhalers = st.number_input("Medical inhalers", step=1000, value=42284)
        emission_data_list = [em_data_grid_electricity, em_data_grid_gas, em_data_bottled_gas, em_data_liquid_fuel, em_data_vehicle_fuel_owned, em_data_business_travel, em_data_anaesthetic_gases, em_data_refrigeration_gases, em_data_waste_management, em_data_medical_inhalers]
        emission_data_list.insert(0, selected_facility)

        st.header("Effect sizes inputs")
        effect_sizes_list = [
            st.slider("Recycling_WasteSegregation_effect", min_value=0.0, max_value=1.0, step=0.1, value=0.90),
            st.slider("SolarSystem_Installation_effect", min_value=0.0, max_value=1.0, step=0.1, value=0.108),
            st.slider("Efficient_Chillers_Upgrade_effect", min_value=0.0, max_value=1.0, step=0.1, value=0.1),
            st.slider("Lighting_Efficiency_effect", min_value=0.0, max_value=1.0, step=0.1, value=0.75),
            st.slider("LowGWP_Refrigerants_effect", min_value=0.0, max_value=1.0, step=0.1, value=0.99),
            st.slider("Hybrid_Car_Use_effect", min_value=0.0, max_value=1.0, step=0.1, value=0.46),
            st.slider("LowGWP_Inhalers_effect", min_value=0.0, max_value=1.0, step=0.1, value=0.999),
            st.slider("LowGWP_AnaestheticGases_effect", min_value=0.0, max_value=1.0, step=0.1, value=0.987),
            st.slider("Staff_Training_Awareness_effect", min_value=0.0, max_value=1.0, step=0.1, value=0.75)
        ]
        effect_sizes_list.insert(0, selected_facility)

        st.header("Implementation costs inputs")
        implementation_costs_list = [
            st.number_input("Recycling & Segregation cost", key="recycling_segregation", step=10000, value=421135),
            st.number_input("Solar Energy Installation cost", key="solar_energy_installation", step=10000, value=50000),
            st.number_input("Efficient Chillers Upgrade cost", key="efficient_chillers_upgrade", step=10000, value=10000),
            st.number_input("LED & Lighting Control cost", key="led_lighting_control", step=10000, value=30000),
            st.number_input("Low-GWP Refrigerants cost", key="low_gwp_refrigerants", step=10000, value=41000),
            st.number_input("Hybrid Vehicles cost", key="hybrid_vehicles", step=10000, value=10000),
            st.number_input("Low-GWP Inhalers cost", key="low_gwp_inhalers", step=10000, value=12000),
            st.number_input("Eco-friendly Anesthetics cost", key="eco_friendly_anesthetics", step=10000, value=39374),
            st.number_input("Emissions Training & Conservation cost", key="emissions_training_conservation", step=10000, value=8000)
        ]
        implementation_costs_list.insert(0, selected_facility)

        st.header("Maintenance costs inputs")
        maintenance_costs_list = [
            st.number_input("Recycling & Segregation cost", step=10000, value=42113),
            st.number_input("Solar Energy Installation cost", step=10000, value=5000),
            st.number_input("Efficient Chillers Upgrade cost", step=10000, value=1000),
            st.number_input("LED & Lighting Control cost", step=10000, value=3000),
            st.number_input("Low-GWP Refrigerants cost", step=10000, value=4100),
            st.number_input("Hybrid Vehicles cost", step=10000, value=1000),
            st.number_input("Low-GWP Inhalers cost", step=10000, value=1200),
            st.number_input("Eco-friendly Anesthetics cost", step=10000, value=3937),
            st.number_input("Emissions Training & Conservation cost", step=10000, value=800)
        ]
        maintenance_costs_list.insert(0, selected_facility)
        generate_data = st.button("Generate Facility Data")
    
    col1, col2 = st.columns(spec=[1, 1], gap="large")
    # Button to update the facility sheet based on the selected facility
    if generate_data:
        update_sheet(file_path, selected_facility, interventions, emission_data_list, effect_sizes_list, implementation_costs_list, maintenance_costs_list)
        # Input data sheet file name (and path if applicable) and read facility code name
        
        input_data_sheet = file_path # MODIFY AS NEEDED
        facility_code = pd.read_excel(input_data_sheet, sheet_name='facility', index_col='Code Name').index[0]
        delete_books("books/carbomica_databook_{}.xlsx".format(facility_code))
        delete_books("books/carbomica_framework_{}.xlsx".format(facility_code))
        delete_books("books/carbomica_progbook_{}.xlsx".format(facility_code))

        # Generate framework, databook and progbook and return facility code name
        generate_books(input_data_sheet, start_year, end_year)

        # Atomica project definition
        P = at.Project(framework = 'books/carbomica_framework_{}.xlsx'.format(facility_code), 
                    databook = 'books/carbomica_databook_{}.xlsx'.format(facility_code), 
                    do_run = False)

        # Projection settings
        P.settings.sim_dt    = 1 # simulation timestep
        P.settings.sim_start = start_year # simulation start year
        P.settings.sim_end   = end_year # simulation end year

        # Load program and define variables for program runs
        progset = P.load_progbook('books/carbomica_progbook_{}.xlsx'.format(facility_code))

        # Set random seed
        np.random.seed(20232212) # MODIFY AS NEEDED

        # Run full coverage scenario            
        with col1:                    
            st.header("Coverage scenario for {}".format(facility_code))
            coverage_scenario(P, progset, start_year, facility_code)
            # Run budget scenario                    
            st.header("Budget scenario for {}".format(facility_code))
            spending = 1e4 # MODIFY AS NEEDED
            budget_scenario(P, progset, start_year, facility_code, spending)
        with col2:
            # Run optimization                    
            budgets = [20e3, 50e3, 100e3] # MODIFY AS NEEDED
            optimization(P, progset, start_year, facility_code, budgets)